from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()
sheet = wb.active

user1 = {'fName': 'Swayam', 
         'lName': 'Panda', 
         'age': 26, 
         'accNo': 2360014534, 
         'ifscCode': 'KKBK0005555',
         'pin': 760001,
         'city': 'Berhampur',
         'isActive': True,
         'designation': 'Python Developer'
         }

user2 = {'fName': 'Avinash', 
         'lName': 'Sahoo', 
         'age': 25, 
         'accNo': 4545623410, 
         'ifscCode': 'PBKK4312424',
         'pin': 500012,
         'city': 'Hyderbad',
         'isActive': True,
         'designation': 'Software Tester'
         }

user3 = {'fName': 'Rajeev', 
         'lName': 'Shukla', 
         'age': 28, 
         'accNo': 4100245679, 
         'ifscCode': 'SBIN454560',
         'pin': 600071,
         'city': 'Chandigarh',
         'isActive': True,
         'designation': 'Data Scientiest'
         }

users = [user1, user2, user3]

rows = []

sheet.append(['First Name', 'Last Name', 'Age', 'Acc No', 'IFSC Code', 'Pin', 'City', 'Active', 'Designation'])
for user in users:
    user_l = list(user.values())
    rows.append(user_l)
    # print(row)
for row in rows:
    sheet.append(row)
    # print("New L:",r)
wb.save("table.xlsx")





    # for key, value in user.items():
    #     row.append(value)
    
# for i in users:
#     l = list(i.values())
#     print(l)

# sheet.append(row)
# wb.save("table.xlsx")

# tab = Table(displayName="Table1", ref="A1:E5")

# style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
#                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
# tab.tableStyleInfo = style



# ------- END ---------#




# from openpyxl import Workbook
# from openpyxl.worksheet.table import Table, TableStyleInfo

# wb = Workbook()
# ws = wb.active

# data = [
#     ['Apples', 10000, 5000, 8000, 6000],
#     ['Pears',   2000, 3000, 4000, 5000],
#     ['Bananas', 6000, 6000, 6500, 6000],
#     ['Oranges',  500,  300,  200,  700],
# ]

# # add column headings. NB. these must be strings
# ws.append(["Fruit", "2011", "2012", "2013", "2014"])
# for row in data:
#     ws.append(row)

# tab = Table(displayName="Table1", ref="A1:E5")

# # Add a default style with striped rows and banded columns
# style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
#                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
# tab.tableStyleInfo = style

# '''
# Table must be added using ws.add_table() method to avoid duplicate names.
# Using this method ensures table name is unque through out defined names and all other table name. 
# '''
# ws.add_table(tab)
# wb.save("table.xlsx")

# wb = Workbook()
# ws = wb.active